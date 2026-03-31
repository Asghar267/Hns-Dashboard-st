#!/usr/bin/env python3
"""
Generate Blink/POS comparison report from SQL Server.

Features:
- all sales vs Blinkco sales vs non-Blinkco sales split
- data quality checks (dedupe, parse, unmatched, mismatch)
- date-stamped outputs under reports/blink/YYYY-MM-DD
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl.styles import Alignment, Font

from modules.database import pool
from modules.blink_reporting import (
    prepare_blink_orders,
    build_split_report,
    build_quality_summary,
)


def _parse_bool(value: str) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _fetch_total_sales(start_date: str, end_exclusive: str, branches: List[int]) -> pd.DataFrame:
    conn = pool.get_connection("candelahns")
    query = f"""
    SELECT
        sh.shop_name,
        COALESCE(e.field_name, 'Online/Unassigned') AS employee_name,
        COUNT(DISTINCT s.sale_id) AS total_transactions_all,
        SUM(s.Nt_amount) AS total_sales_all
    FROM tblSales s
    LEFT JOIN tblDefShopEmployees e ON s.employee_id = e.shop_employee_id
    LEFT JOIN tblDefShops sh ON s.shop_id = sh.shop_id
    WHERE s.sale_date >= ?
      AND s.sale_date < ?
      AND s.shop_id IN ({",".join(["?"] * len(branches))})
    GROUP BY sh.shop_name, COALESCE(e.field_name, 'Online/Unassigned')
    """
    return pd.read_sql(query, conn, params=[start_date, end_exclusive] + branches)


def _fetch_blinkco_sales(start_date: str, end_exclusive: str, branches: List[int]) -> pd.DataFrame:
    conn = pool.get_connection("candelahns")
    query = f"""
    SELECT
        s.sale_id,
        sh.shop_name,
        COALESCE(e.field_name, 'Online/Unassigned') AS employee_name,
        s.Nt_amount AS total_sale,
        s.external_ref_id
    FROM tblSales s
    LEFT JOIN tblDefShopEmployees e ON s.employee_id = e.shop_employee_id
    LEFT JOIN tblDefShops sh ON s.shop_id = sh.shop_id
    WHERE s.sale_date >= ?
      AND s.sale_date < ?
      AND s.shop_id IN ({",".join(["?"] * len(branches))})
      AND s.external_ref_type = 'Blinkco order'
    """
    return pd.read_sql(query, conn, params=[start_date, end_exclusive] + branches)


def _fetch_blink_raw(start_date: str, end_exclusive: str) -> pd.DataFrame:
    conn = pool.get_connection("candelahns")
    query = """
    SELECT BlinkOrderId, OrderJson, CreatedAt
    FROM tblInitialRawBlinkOrder
    WHERE CreatedAt >= ? AND CreatedAt < ?
    """
    return pd.read_sql(query, conn, params=[start_date, end_exclusive])


def _build_outputs(
    start_date: str,
    end_exclusive: str,
    comm_rate: float,
    branches: List[int],
    include_unassigned: bool,
    top_n: int,
):
    df_total = _fetch_total_sales(start_date, end_exclusive, branches)
    df_blinkco_tx = _fetch_blinkco_sales(start_date, end_exclusive, branches)
    df_raw = _fetch_blink_raw(start_date, end_exclusive)

    raw_rows = len(df_raw)
    df_blink = prepare_blink_orders(df_raw)
    deduped_rows = len(df_blink)

    if not df_blinkco_tx.empty:
        df_blinkco_tx["total_sale"] = pd.to_numeric(df_blinkco_tx["total_sale"], errors="coerce").fillna(0.0)
        df_blinkco_tx = df_blinkco_tx.merge(
            df_blink,
            left_on="external_ref_id",
            right_on="BlinkOrderId",
            how="left",
        )
        df_blinkco_tx["Indoge_total_price"] = pd.to_numeric(df_blinkco_tx["Indoge_total_price"], errors="coerce").fillna(0.0)
        df_blinkco_tx["BlinkOrderId"] = df_blinkco_tx["BlinkOrderId"].fillna("-")
        df_blinkco_tx["json_parse_ok"] = df_blinkco_tx["json_parse_ok"].fillna(False)
        df_blinkco_tx["difference"] = df_blinkco_tx["Indoge_total_price"] - df_blinkco_tx["total_sale"]
        df_blinkco_tx["has_blink_order"] = df_blinkco_tx["BlinkOrderId"].astype(str).ne("-")
        df_blinkco_tx["blink_mismatch_flag"] = df_blinkco_tx["has_blink_order"] & (df_blinkco_tx["difference"].abs() > 1.0)
        df_blinkco_tx["Candelahns_commission"] = df_blinkco_tx["total_sale"] * (comm_rate / 100.0)
        df_blinkco_tx["Indoge_commission"] = df_blinkco_tx["Indoge_total_price"] * (comm_rate / 100.0)

    blinkco_summary = (
        df_blinkco_tx.groupby(["employee_name", "shop_name"], as_index=False)
        .agg(
            total_sales_blinkco=("total_sale", "sum"),
            total_transactions_blinkco=("sale_id", "count"),
        )
        if not df_blinkco_tx.empty
        else pd.DataFrame(columns=["employee_name", "shop_name", "total_sales_blinkco", "total_transactions_blinkco"])
    )

    split_report = build_split_report(df_total, blinkco_summary, comm_rate)
    if not include_unassigned and not split_report.empty:
        split_report = split_report[split_report["employee_name"].astype(str).str.strip().str.lower() != "online/unassigned"]
        df_blinkco_tx = df_blinkco_tx[df_blinkco_tx["employee_name"].astype(str).str.strip().str.lower() != "online/unassigned"]

    split_report = split_report.sort_values("total_sales_all", ascending=False).reset_index(drop=True)

    totals_summary = pd.DataFrame(
        [
            {
                "scope": "ALL_ORDER_TYPES",
                "total_transactions": int(split_report["total_transactions_all"].sum()) if not split_report.empty else 0,
                "total_sales": float(split_report["total_sales_all"].sum()) if not split_report.empty else 0.0,
                "commission": float(split_report["commission_total_sales"].sum()) if not split_report.empty else 0.0,
            },
            {
                "scope": "BLINKCO_ONLY",
                "total_transactions": int(split_report["total_transactions_blinkco"].sum()) if not split_report.empty else 0,
                "total_sales": float(split_report["total_sales_blinkco"].sum()) if not split_report.empty else 0.0,
                "commission": float(split_report["commission_blinkco_sales"].sum()) if not split_report.empty else 0.0,
            },
            {
                "scope": "WITHOUT_BLINKCO",
                "total_transactions": int(split_report["total_transactions_without_blinkco"].sum()) if not split_report.empty else 0,
                "total_sales": float(split_report["total_sales_without_blinkco"].sum()) if not split_report.empty else 0.0,
                "commission": float(split_report["commission_without_blinkco_sales"].sum()) if not split_report.empty else 0.0,
            },
        ]
    )
    all_sales = totals_summary.loc[totals_summary["scope"] == "ALL_ORDER_TYPES", "total_sales"].iloc[0] if not totals_summary.empty else 0.0
    totals_summary["sales_pct_of_all"] = (totals_summary["total_sales"] / all_sales * 100).round(2) if all_sales else 0.0

    quality_summary = build_quality_summary(df_blinkco_tx, raw_rows=raw_rows, deduped_rows=deduped_rows)

    recon_checks = pd.DataFrame(
        [
            {
                "check": "split_total_matches_source_total",
                "lhs": float(split_report["total_sales_all"].sum()) if not split_report.empty else 0.0,
                "rhs": float(df_total["total_sales_all"].sum()) if not df_total.empty else 0.0,
            },
            {
                "check": "split_blink_matches_tx_blink",
                "lhs": float(split_report["total_sales_blinkco"].sum()) if not split_report.empty else 0.0,
                "rhs": float(df_blinkco_tx["total_sale"].sum()) if not df_blinkco_tx.empty else 0.0,
            },
        ]
    )
    recon_checks["abs_diff"] = (recon_checks["lhs"] - recon_checks["rhs"]).abs().round(4)
    recon_checks["status"] = recon_checks["abs_diff"].apply(lambda x: "PASS" if x <= 1 else "WARN")

    top_movers = split_report.head(max(top_n, 1)).copy()
    return totals_summary, split_report, df_blinkco_tx, quality_summary, recon_checks, top_movers


def _format_excel(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.auto_filter.ref = ws.dimensions

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if "pct" in str(ws.cell(row=1, column=cell.column).value).lower():
                        cell.number_format = "0.00"
                    else:
                        cell.number_format = "#,##0.00"
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Blink/POS comparison report.")
    parser.add_argument("--start-date", default="2026-01-25")
    parser.add_argument("--end-exclusive", default="2026-02-26", help="Upper bound, excluded")
    parser.add_argument("--comm-rate", type=float, default=2.0)
    parser.add_argument("--branches", default="2,3,4,6,8,10,14")
    parser.add_argument("--include-unassigned", default="true")
    parser.add_argument("--top-n", type=int, default=20)
    parser.add_argument("--output-root", default="reports/blink")
    parser.add_argument("--markdown-brief", default="true")
    args = parser.parse_args()

    branches = [int(x.strip()) for x in args.branches.split(",") if x.strip()]
    include_unassigned = _parse_bool(args.include_unassigned)
    markdown_brief = _parse_bool(args.markdown_brief)

    totals_summary, split_report, df_blinkco_tx, quality_summary, recon_checks, top_movers = _build_outputs(
        args.start_date,
        args.end_exclusive,
        args.comm_rate,
        branches,
        include_unassigned,
        args.top_n,
    )

    run_stamp = datetime.now().strftime("%Y-%m-%d")
    out_dir = Path(args.output_root) / run_stamp
    out_dir.mkdir(parents=True, exist_ok=True)

    totals_summary.to_csv(out_dir / "blink_totals_summary.csv", index=False)
    split_report.to_csv(out_dir / "blink_employee_branch_split.csv", index=False)
    df_blinkco_tx.to_csv(out_dir / "blink_blinkco_transactions.csv", index=False)
    quality_summary.to_csv(out_dir / "blink_data_quality_checks.csv", index=False)
    recon_checks.to_csv(out_dir / "blink_reconciliation_checks.csv", index=False)

    xlsx_path = out_dir / "blink_sales_report_comparison.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        totals_summary.to_excel(writer, sheet_name="summary", index=False)
        split_report.to_excel(writer, sheet_name="employee_branch_split", index=False)
        df_blinkco_tx.to_excel(writer, sheet_name="blinkco_transactions", index=False)
        quality_summary.to_excel(writer, sheet_name="data_quality_checks", index=False)
        recon_checks.to_excel(writer, sheet_name="reconciliation_checks", index=False)
    _format_excel(xlsx_path)

    md_lines = [
        "# Blink Sales Comparison Report",
        f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**Date Range:** {args.start_date} to {args.end_exclusive} (exclusive)",
        f"**Commission Rate:** {args.comm_rate}%",
        f"**Include Unassigned:** {include_unassigned}",
        "",
        "## Summary",
        totals_summary.to_markdown(index=False),
        "",
        "## Reconciliation Checks",
        recon_checks.to_markdown(index=False),
        "",
        f"## Top {max(args.top_n,1)} Movers (by total_sales_all)",
        top_movers.to_markdown(index=False) if not top_movers.empty else "_No data_",
        "",
        "## Data Quality Checks",
        quality_summary.to_markdown(index=False),
        "",
    ]
    if markdown_brief:
        all_row = totals_summary[totals_summary["scope"] == "ALL_ORDER_TYPES"].iloc[0]
        blink_row = totals_summary[totals_summary["scope"] == "BLINKCO_ONLY"].iloc[0]
        md_lines.extend(
            [
                "## Executive Brief",
                f"- Total sales: {all_row['total_sales']:,.2f}",
                f"- Blinkco sales: {blink_row['total_sales']:,.2f} ({blink_row['sales_pct_of_all']:.2f}% of total)",
                f"- Blinkco commission: {blink_row['commission']:,.2f}",
                "",
            ]
        )

    md_path = out_dir / "blink_sales_report_comparison.md"
    md_path.write_text("\n".join(md_lines), encoding="utf-8")

    print(totals_summary.to_string(index=False))
    print(f"\nSaved report artifacts in: {out_dir}")


if __name__ == "__main__":
    main()
